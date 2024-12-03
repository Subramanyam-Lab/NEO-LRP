from flp import createFlp
from vrp import createVRP
from solver_cvrp import dataCvrp
import os
import openpyxl
from dataparse import create_data
from datetime import datetime
import logging
import sys

# Set up logging directory
log_dir = "log_files/ortools"
os.makedirs(log_dir, exist_ok=True)

# Directory containing the prodhon dataset
directory_path = "NEO-LRP/prodhon_dataset"  

# Excel file to store the results
existing_excel_file = "/storage/group/azs7266/default/wzk5140/MLforVRP/Codes/vpr_model/NEO-LRP/results/flp/flp.xlsx" 

bks_dict = {
    "coord20-5-1.dat": 54793,
    "coord20-5-1b.dat": 39104,
    "coord20-5-2.dat": 48908,
    "coord20-5-2b.dat": 37542,
    "coord50-5-1.dat": 90111,
    "coord50-5-1b.dat": 63242,
    "coord50-5-2.dat": 88293,
    "coord50-5-2b.dat": 67308,
    "coord50-5-2bBIS.dat": 51822,
    "coord50-5-2BIS.dat": 84055,
    "coord50-5-3.dat": 86203,
    "coord50-5-3b.dat": 61830,
    "coord100-5-1.dat": 274814,
    "coord100-5-1b.dat": 213568,
    "coord100-5-2.dat": 193671,
    "coord100-5-2b.dat": 157095,
    "coord100-5-3.dat": 200079,
    "coord100-5-3b.dat": 152441,
    "coord100-10-1.dat": 287661,
    "coord100-10-1b.dat": 230989,
    "coord100-10-2.dat": 243590,
    "coord100-10-2b.dat": 203988,
    "coord100-10-3.dat": 250882,
    "coord100-10-3b.dat": 203114,
    "coord200-10-1.dat": 474850,
    "coord200-10-1b.dat": 375177,
    "coord200-10-2.dat": 448077,
    "coord200-10-2b.dat": 373696,
    "coord200-10-3.dat": 469433,
    "coord200-10-3b.dat": 362320
}

headers = [
    "Instance",
    "FLP Cost",
    "VRP Route Total Cost",
    "VRP LRP Cost",
    "FLP Execution Time (s)",
    "VRP OR Total Execution Time (s)",
    "VRP OR Execution Time per Order (s)",
    "VRP Easy Total VRP Cost",
    "VRP Easy Total LRP Cost",
    "VRP Easy Total Execution Time (s)",
    "VRP Easy Execution Time per Operation (s)",
    "Gap"
]

# Function to set headers if not already present
def set_headers(ws, headers):
    existing_headers = [cell.value for cell in ws[1]]
    if not existing_headers or existing_headers != headers:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)

# Load or create the Excel workbook
if not os.path.exists(existing_excel_file):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "FLP_VRP_Results"
    set_headers(worksheet, headers)
    workbook.save(existing_excel_file)
    print(f"Created new Excel file with headers at {existing_excel_file}")
else:
    workbook = openpyxl.load_workbook(existing_excel_file)
    worksheet = workbook.active
    set_headers(worksheet, headers)

# Keep track of processed instances to avoid duplicates
processed_instances = set()
for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    instance_name = row[0]
    if instance_name is not None:
        processed_instances.add(instance_name)

# Main processing loop
for filename in bks_dict.keys():
    if filename in processed_instances:
        print(f"Instance {filename} already processed. Skipping.")
        continue

    # Initialize lists to collect metrics for averaging
    flp_cost_list = []
    or_vrp_route_total_cost_list = []
    or_vrp_lrp_cost_list = []
    flp_execution_time_list = []
    vrp_or_total_execution_time_list = []
    vrp_or_execution_time_per_order_list = []
    vrp_easy_total_vrp_cost_list = []
    vrp_easy_total_lrp_cost_list = []
    vrp_easy_total_execution_time_list = []
    vrp_easy_execution_time_per_operation_list = []
    gap_list = []

    file_path = os.path.join(directory_path, filename)
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        continue

    print("Working on:", file_path)

    for run_index in range(5):
        print(f"Run {run_index + 1} for instance {filename}")

        # Create the log file name based on the input file's name
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        log_filename = f"{os.path.splitext(filename)[0]}_{current_time}_run_{run_index}.log"

        # Configure logging
        logger = logging.getLogger()
        logger.setLevel(logging.INFO)
        # Remove all handlers associated with the root logger object
        for handler in logger.handlers[:]:
            logger.removeHandler(handler)
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(os.path.join(log_dir, log_filename)),
                logging.StreamHandler(sys.stdout)
            ]
        )

        # Process the data
        ans = create_data(file_path)
        flp_solver = createFlp(ans)
        flp_results = flp_solver.flp()
        logging.debug("FLP Results: %s", flp_results)
        od = len(flp_results[1])

        # FLP Execution Time
        flp_execution_time = flp_results[5]

        # VRP OR execution
        or_st = datetime.now()
        vrp_solver = createVRP(ans, flp_results)
        or_vrp_results = vrp_solver.runVRP()
        logging.debug("VRP Results: %s", or_vrp_results)
        or_ed = datetime.now()
        tot_or_exec = (or_ed - or_st).total_seconds()
        or_exec = tot_or_exec / od if od != 0 else 0

        # VRP Easy execution
        ve_st = datetime.now()
        print("Running vrpeasy function")
        vrpeasy_solver = dataCvrp(ans, flp_results)
        vrp_easy_results = vrpeasy_solver.runVRPeasy()
        logging.info("VRP Results: %s", vrp_easy_results)
        ve_ed = datetime.now()
        tot_ve_exec = (ve_ed - ve_st).total_seconds()
        ve_exec = tot_ve_exec / od if od != 0 else 0

        # Calculate Gap
        bks = bks_dict.get(filename)
        vrp_easy_total_lrp_cost = vrp_easy_results[0]
        if bks is not None and bks != 0:
            gap = (abs(bks - vrp_easy_total_lrp_cost) / bks) * 100
            gap = round(gap, 2)
        else:
            gap = None

        # Append metrics to lists
        flp_cost_list.append(flp_results[0])
        or_vrp_route_total_cost_list.append(or_vrp_results[0])
        or_vrp_lrp_cost_list.append(or_vrp_results[1])
        flp_execution_time_list.append(flp_execution_time)
        vrp_or_total_execution_time_list.append(tot_or_exec)
        vrp_or_execution_time_per_order_list.append(or_exec)
        vrp_easy_total_vrp_cost_list.append(vrp_easy_results[1])
        vrp_easy_total_lrp_cost_list.append(vrp_easy_total_lrp_cost)
        vrp_easy_total_execution_time_list.append(tot_ve_exec)
        vrp_easy_execution_time_per_operation_list.append(ve_exec)
        gap_list.append(gap if gap is not None else 0)

        # Close the log file handlers
        for handler in logging.root.handlers[:]:
            handler.close()
            logging.root.removeHandler(handler)

    # Compute averages
    avg_flp_cost = sum(flp_cost_list) / len(flp_cost_list)
    avg_or_vrp_route_total_cost = sum(or_vrp_route_total_cost_list) / len(or_vrp_route_total_cost_list)
    avg_or_vrp_lrp_cost = sum(or_vrp_lrp_cost_list) / len(or_vrp_lrp_cost_list)
    avg_flp_execution_time = sum(flp_execution_time_list) / len(flp_execution_time_list)
    avg_vrp_or_total_execution_time = sum(vrp_or_total_execution_time_list) / len(vrp_or_total_execution_time_list)
    avg_vrp_or_execution_time_per_order = sum(vrp_or_execution_time_per_order_list) / len(vrp_or_execution_time_per_order_list)
    avg_vrp_easy_total_vrp_cost = sum(vrp_easy_total_vrp_cost_list) / len(vrp_easy_total_vrp_cost_list)
    avg_vrp_easy_total_lrp_cost = sum(vrp_easy_total_lrp_cost_list) / len(vrp_easy_total_lrp_cost_list)
    avg_vrp_easy_total_execution_time = sum(vrp_easy_total_execution_time_list) / len(vrp_easy_total_execution_time_list)
    avg_vrp_easy_execution_time_per_operation = sum(vrp_easy_execution_time_per_operation_list) / len(vrp_easy_execution_time_per_operation_list)
    avg_gap = sum(gap_list) / len([g for g in gap_list if g != 0]) if any(gap_list) else "N/A"

    # Create a new row
    new_row = [
        os.path.basename(file_path),
        avg_flp_cost,
        avg_or_vrp_route_total_cost,
        avg_or_vrp_lrp_cost,
        avg_flp_execution_time,
        avg_vrp_or_total_execution_time,
        avg_vrp_or_execution_time_per_order,
        avg_vrp_easy_total_vrp_cost,
        avg_vrp_easy_total_lrp_cost,
        avg_vrp_easy_total_execution_time,
        avg_vrp_easy_execution_time_per_operation,
        avg_gap
    ]

    # Append the new row to the worksheet
    worksheet.append(new_row)

    # Save the modified Excel file
    workbook.save(existing_excel_file)

    # Add the instance to the processed instances set
    processed_instances.add(filename)
