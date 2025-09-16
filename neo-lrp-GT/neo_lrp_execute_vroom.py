"""
Neural Embedded Mixed-Integer Optimization for Location-Routing Problems
Main execution script using Graph Transformer and VROOM solver
"""

from neural_embedded_model import createLRP, write_to_txt_cvrplib_format
from solver_cvrp_vroom import dataCvrp

import os
import shutil
import argparse
import openpyxl
from dataparse import create_data
from datetime import datetime
import logging
import sys
import json
from openpyxl import Workbook, load_workbook

# Configuration parameters
BFS = "solutions"
phi_loc = "../pre_trained_models/graph_transformer.pth"
existing_excel_file = "results/neo_lrp_gt_results.xlsx"
sheet_name = "results"
fi_mode_input = "dynamic"

# Create necessary directories
log_dir = "log_files/mip_nn"
os.makedirs(log_dir, exist_ok=True)
os.makedirs("results", exist_ok=True)
directory_path = "../prodhon_dataset"
 
if os.path.exists(BFS):
    print(f"[INFO] Removing old BFS folder: {BFS}")
    shutil.rmtree(BFS)
os.makedirs(BFS, exist_ok=True)
print(f"[INFO] Created fresh BFS folder: {BFS}")


try:
    # Remove existing results file if it exists
    if os.path.exists(existing_excel_file):
        print(f"Existing results file found: {existing_excel_file}. Deleting it.")
        os.remove(existing_excel_file)

    # Create new results file
    workbook = Workbook()
    workbook.save(existing_excel_file)
    print(f"New results file created: {existing_excel_file}")

except Exception as e:
    print(f"Error while handling Excel file: {e}")
    print("Recreating the Excel file...")
    # Delete and recreate file
    if os.path.exists(existing_excel_file):
        os.remove(existing_excel_file)
    workbook = Workbook()
    workbook.save(existing_excel_file)
    print(f"New file created: {existing_excel_file}")

if sheet_name not in workbook.sheetnames:
    workbook.create_sheet(sheet_name)
worksheet = workbook[sheet_name]


headings = [
    "Instance", "FLP", "VRP", "LRP(MIP+NN)", "NumRoutes_OptSol",
    "LA time", "VRPSolverEasy computed VRP cost",
    "actual LRP cost(using VRPSolverEasy)",
    "Solver time", "BKS",
    "Optimization_gap_optsol", "Prediction_gap"
]

if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet.cell(1, 1).value is None:
    # Add headings to the first row
    for col, heading in enumerate(headings, start=1):
        worksheet.cell(row=1, column=col, value=heading)

def has_customers(instance_file_path):
    """
    Check if a VRP instance has any customers assigned.
    
    Args:
        instance_file_path (str): Path to the VRP instance file
        
    Returns:
        bool: True if the instance has customers, False otherwise
        
    Raises:
        ValueError: If required sections are missing from the file
    """
    with open(instance_file_path, 'r') as file:
        lines = file.readlines()
    try:
        demand_section_index = lines.index("DEMAND_SECTION\n")
        depot_section_index = lines.index("DEPOT_SECTION\n")
    except ValueError as e:
        raise ValueError(f"Section missing in file {instance_file_path}: {e}")
    demand_lines = lines[demand_section_index + 1:depot_section_index]
    # Remove depot demand (assumed to be the first line)
    customer_demands = demand_lines[1:]
    return len(customer_demands) > 0

# Best Known Solutions (BKS) for benchmark instances
BKS_DICT = {
    "coord20-5-1.dat": 54793,
    "coord20-5-1b.dat": 39104,
    "coord20-5-2.dat": 48908,
    "coord20-5-2b.dat": 37542,
    "coord50-5-1.dat": 90111,
    "coord50-5-1b.dat": 63242,
    "coord50-5-2.dat": 88293,
    "coord50-5-2b.dat": 67308,
    "coord50-5-2bBIS.dat": 51822,
    "coord50-5-2BIS.dat": 84055,
    "coord50-5-3.dat": 86203,
    "coord50-5-3b.dat": 61830,
    "coord100-5-1.dat": 274814,
    "coord100-5-1b.dat": 213568,
    "coord100-5-2.dat": 193671,
    "coord100-5-2b.dat": 157095,
    "coord100-5-3.dat": 200079,
    "coord100-5-3b.dat": 152441,
    "coord100-10-1.dat": 287661,
    "coord100-10-1b.dat": 230989,
    "coord100-10-2.dat": 243590,
    "coord100-10-2b.dat": 203988,
    "coord100-10-3.dat": 250882,
    "coord100-10-3b.dat": 203114,
    "coord200-10-1.dat": 474850,
    "coord200-10-1b.dat": 375177,
    "coord200-10-2.dat": 448077,
    "coord200-10-2b.dat": 373696,
    "coord200-10-3.dat": 469433,
    "coord200-10-3b.dat": 362320
}

# Track processed instances to avoid duplicates
processed_instances = set()
for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    instance_name = row[0]
    if instance_name is not None:
        processed_instances.add(instance_name)
        
# Main execution loop for each benchmark instance
for filename in BKS_DICT.keys(): 
    # Skip if already processed (uncomment if needed)
    # if filename in processed_instances:
    #     print(f"Instance {filename} already processed. Skipping.")
    #     continue

    # Initialize lists to collect metrics for averaging across multiple runs
    flp_cost_list = []
    vrp_cost_list = []
    lrp_cost_list = []
    vrp_routes_optsol_list = []
    lrp_exec_list = []
    warmstart_time_list = []
    nn_model_time_list = []
    vrp_easy_vrp_cost_list = []
    actual_lrp_cost_list = []
    ve_exec_list = []
    tot_ve_exec_list = []
    vrp_solver_easy_model_solve_time_list = []
    gap_list = []
    gap_vrp_perc_list = []
        
    file_path = os.path.join(directory_path, filename)
    if os.path.exists(file_path):
        print("Working on:", file_path)
    else:
        print("File not found:", file_path)
        break

    # Run each instance multiple times for statistical significance
    for run_index in range(5):
        print(f"Run {run_index + 1} for instance {filename}")

        # Define per-run subdirectory for storing intermediate results
        instance_subdir_run = os.path.join(BFS, os.path.splitext(filename)[0], f"run_{run_index}")
        os.makedirs(instance_subdir_run, exist_ok=True)

        # Prepare logging with timestamp
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        log_filename = f"{os.path.splitext(filename)[0]}_{current_time}_run_{run_index}.log"

        print(f'\n\n Working on file :{file_path}')
        ans = create_data(file_path)  # Process the data

        # Execute neural embedded LRP solver
        lrp_st = datetime.now()
        lrp_solver = createLRP(ans)
        lrp_result = lrp_solver.model(
            file_path,
            log_filename,
            instance_subdir_run,
            phi_loc,
            fi_mode=fi_mode_input,
            fixed_fi_value=1000.0
        )
        lrp_ed = datetime.now()

        warmstart_time = lrp_result[4]
        pred_costs_dict = lrp_result[5]  # {depot_id: NN predicted cost}

        # Process facility assignment results
        flp_dict = {}
        for j in range(len(lrp_result[0])):  # lrp_result[0] : y[j] (facility assignment)
            if lrp_result[0][j] > 0.5:
                ls = []
                for i in range(len(lrp_result[1][j])):
                    if lrp_result[1][j][i] > 0.5:
                        ls.append(i)
                flp_dict[j] = ls

        # Build routing data structures for VROOM solver
        rout_dist = {}
        fac_cust_dem = {}
        cust_dem_fac = {}
        for f in flp_dict:
            ls1 = []
            ls2 = []
            dem_sum = 0
            for c in flp_dict[f]:
                ls1.append(ans[3][c])  # Customer coordinates
                dem_sum += ans[6][c]   # Customer demand
                ls2.append(ans[6][c])
            ls1.insert(0, ans[2][f])   # Insert depot coordinates at the beginning
            ls2.insert(0, 0)           # Insert depot demand (0) at the beginning
            rout_dist[f] = ls1
            fac_cust_dem[f] = dem_sum
            cust_dem_fac[f] = ls2
        ass_result = [lrp_result[2], flp_dict, rout_dist, fac_cust_dem, cust_dem_fac]
        
        # Execute VROOM solver for exact VRP solution
        ve_st = datetime.now()
        vrpeasy_solver = dataCvrp(ans, ass_result)
        vrp_easy_results = vrpeasy_solver.runVRPeasy()
        ve_ed = datetime.now()

        # Calculate performance metrics
        num_open_depots = len(flp_dict)

        lrp_exec = ((lrp_ed - lrp_st).total_seconds()) / num_open_depots
        warmstart_time = warmstart_time / num_open_depots
        tot_ve_exec = (ve_ed - ve_st).total_seconds()
        ve_exec = tot_ve_exec / num_open_depots

        instance_name = os.path.basename(file_path)
        bks = BKS_DICT.get(instance_name)

        flp_cost = lrp_result[2]
        vrp_cost = lrp_result[3]
        lrp_cost = flp_cost + vrp_cost
        actual_lrp_cost = vrp_easy_results[0]
        vrp_easy_vrp_cost = vrp_easy_results[1]
        vrp_routes_optsol = sum(vrp_easy_results[3])

        # Calculate optimization gap compared to BKS
        if bks is not None and bks != 0:
            gap = (abs(bks - actual_lrp_cost) / bks) * 100
            gap = round(gap, 2)
        else:
            gap = "N/A"

        # Calculate prediction gap for neural network
        actual_costs_list = vrp_easy_results[2]  # Actual VRP cost per depot
        open_depots = list(flp_dict.keys())      # Open depot IDs, same order as pred_costs_dict

        per_depot_gaps = []
        for k, depot_id in enumerate(open_depots):
            true_c = actual_costs_list[k]
            pred_c = pred_costs_dict.get(depot_id, 0.0)
            if true_c:  # Avoid division by zero
                per_depot_gaps.append(abs(pred_c - true_c) / true_c * 100)

        gap_vrp_perc = (
            round(sum(per_depot_gaps) / len(per_depot_gaps), 2)
            if per_depot_gaps else "N/A"
        )

        vrp_solver_easy_model_solve_time = vrp_easy_results[7]

        # Append metrics to lists
        flp_cost_list.append(flp_cost)
        vrp_cost_list.append(vrp_cost)
        lrp_cost_list.append(lrp_cost)
        vrp_routes_optsol_list.append(vrp_routes_optsol)
        lrp_exec_list.append(lrp_exec)
        warmstart_time_list.append(warmstart_time)
        # nn_model_time_list.append(nn_model_time)
        vrp_easy_vrp_cost_list.append(vrp_easy_vrp_cost)
        actual_lrp_cost_list.append(actual_lrp_cost)
        ve_exec_list.append(ve_exec)
        tot_ve_exec_list.append(tot_ve_exec)
        vrp_solver_easy_model_solve_time_list.append(vrp_solver_easy_model_solve_time)
        gap_list.append(gap if gap != "N/A" else 0)
        gap_vrp_perc_list.append(gap_vrp_perc if gap_vrp_perc != "N/A" else 0)

    # Compute average metrics across all runs
    avg_flp_cost = sum(flp_cost_list) / len(flp_cost_list)
    avg_vrp_cost = sum(vrp_cost_list) / len(vrp_cost_list)
    avg_lrp_cost = sum(lrp_cost_list) / len(lrp_cost_list)
    avg_vrp_routes_optsol = sum(vrp_routes_optsol_list) / len(vrp_routes_optsol_list)
    avg_lrp_exec = sum(lrp_exec_list) / len(lrp_exec_list)
    avg_warmstart_time = sum(warmstart_time_list) / len(warmstart_time_list)
    avg_vrp_easy_vrp_cost = sum(vrp_easy_vrp_cost_list) / len(vrp_easy_vrp_cost_list)
    avg_actual_lrp_cost = sum(actual_lrp_cost_list) / len(actual_lrp_cost_list)
    avg_ve_exec = sum(ve_exec_list) / len(ve_exec_list)
    avg_tot_ve_exec = sum(tot_ve_exec_list) / len(tot_ve_exec_list)
    avg_vrp_solver_easy_model_solve_time = sum(vrp_solver_easy_model_solve_time_list) / len(vrp_solver_easy_model_solve_time_list)
    avg_gap = sum(gap_list) / len(gap_list) if gap_list else "N/A"
    avg_gap_vrp_perc = sum(gap_vrp_perc_list) / len(gap_vrp_perc_list) if gap_vrp_perc_list else "N/A"

    # Build result row for Excel output
    new_row = [
        os.path.basename(file_path),  # Instance name
        avg_flp_cost,                 # Average FLP cost
        avg_vrp_cost,                 # Average VRP cost
        avg_lrp_cost,                 # Average LRP cost (FLP + VRP)
        avg_vrp_routes_optsol,        # Average number of routes in optimal solution
        avg_lrp_exec,                 # Average LRP solver execution time
        avg_vrp_easy_vrp_cost,        # Average VROOM computed VRP cost
        avg_actual_lrp_cost,          # Average actual LRP cost using VROOM
        avg_tot_ve_exec,              # Average VROOM solver execution time
        bks,                          # Best known solution
        avg_gap,                      # Average optimization gap
        avg_gap_vrp_perc,             # Average prediction gap
    ]

    # Save results to Excel file
    worksheet.append(new_row)
    workbook.save(existing_excel_file)