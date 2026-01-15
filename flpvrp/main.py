"""
Main entry point for solving benchmark CLRP instances using the FLP-VRP appraoch.
First solves CFLP (using SCIP) to determine depot locations then VROOM to compute vehicle routes by solving CVRP.
Saves results to Excel and JSON files for analysis.
"""

from flp_solver import createFlp
from vrp_solver_vroom import dataCvrp
import os
import openpyxl
from dataparse import create_data
from datetime import datetime
import logging
import sys
import json
from pathlib import Path

SOLN_DIR = Path("results/solutions_flpvrp")
SOLN_DIR.mkdir(parents=True, exist_ok=True)

def dump_json(path: Path, payload: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

log_dir = "log_files"
os.makedirs(log_dir, exist_ok=True)

directory_path = "NEO-LRP/benchmark_instances/P_prodhon"
existing_excel_file = "NEO-LRP/flpvrp/lp_results.xlsx" 

bks_dict = {
    "coord20-5-1.dat": 54793,
    "coord20-5-1b.dat": 39104,
    "coord20-5-2.dat": 48908,
    "coord20-5-2b.dat": 37542,
    "coord50-5-1.dat": 90111,
    "coord50-5-1b.dat": 63242,
    "coord50-5-2.dat": 88293,
    "coord50-5-2b.dat": 67308,
    "coord50-5-2bBIS.dat": 51822,
    "coord50-5-2BIS.dat": 84055,
    "coord50-5-3.dat": 86203,
    "coord50-5-3b.dat": 61830,
    "coord100-5-1.dat": 274814,
    "coord100-5-1b.dat": 213568,
    "coord100-5-2.dat": 193671,
    "coord100-5-2b.dat": 157095,
    "coord100-5-3.dat": 200079,
    "coord100-5-3b.dat": 152441,
    "coord100-10-1.dat": 287661,
    "coord100-10-1b.dat": 230989,
    "coord100-10-2.dat": 243590,
    "coord100-10-2b.dat": 203988,
    "coord100-10-3.dat": 250882,
    "coord100-10-3b.dat": 203114,
    "coord200-10-1.dat": 474702,
    "coord200-10-1b.dat": 375177,
    "coord200-10-2.dat": 448077,
    "coord200-10-2b.dat": 373696,
    "coord200-10-3.dat": 469433,
    "coord200-10-3b.dat": 362320
}

headers = [
    "Instance",
    "FLP Cost",
    "FLP Surrogate Cost (assignment)",
    "FLP Time (s)",
    "VROOM VRP Cost (travel+fixed)",
    "VROOM LRP Cost (FLP+VRP)",
    "VROOM VRP Time (s)",
    "VROOM VRP Time / Depot (s)",
    "Gap to BKS (vs VROOM LRP) (%)",
    "Gap Sugg (vs VROOM VRP) (%)"
]


def set_headers(ws, headers):
    existing_headers = [cell.value for cell in ws[1]]
    if not existing_headers or existing_headers != headers:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)

if not os.path.exists(existing_excel_file):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "FLP_VRP_Results"
    set_headers(worksheet, headers)
    workbook.save(existing_excel_file)
    print(f"Created new Excel file with headers at {existing_excel_file}")
else:
    workbook = openpyxl.load_workbook(existing_excel_file)
    worksheet = workbook.active
    set_headers(worksheet, headers)

processed_instances = set()
for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    instance_name = row[0]
    if instance_name is not None:
        processed_instances.add(instance_name)

for filename in bks_dict.keys():
    if filename in processed_instances:
        print(f"Instance {filename} already processed. Skipping.")
        continue

    flp_cost_list = []
    flp_execution_time_list = []
    VROOM_total_vrp_cost_list = []
    VROOM_total_lrp_cost_list = []
    VROOM_total_execution_time_list = []
    VROOM_execution_time_per_operation_list = []
    gap_list = []
    sur_cost_list = []
    gap_sugg_list = []

    file_path = os.path.join(directory_path, filename)
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        continue

    print("Working on:", file_path)

    for run_index in range(1):
        print(f"Run {run_index + 1} for instance {filename}")

        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        log_filename = f"{os.path.splitext(filename)[0]}_{current_time}_run_{run_index}.log"

        logger = logging.getLogger()
        logger.setLevel(logging.INFO)
        for handler in logger.handlers[:]:
            logger.removeHandler(handler)
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(os.path.join(log_dir, log_filename)),
                logging.StreamHandler(sys.stdout)
            ]
        )

        ans = create_data(file_path)
        flp_solver = createFlp(ans)
        flp_results = flp_solver.flp()
        sur_cost = flp_results[6] 

        if not hasattr(flp_solver, "last_solution") or flp_solver.last_solution is None:
            print(f"[WARN] No FLP JSON payload for {filename}, run {run_index+1}")
        else:
            instance_stem = os.path.splitext(filename)[0]
            flp_json = SOLN_DIR / instance_stem / "FLPVRP" / f"run_{run_index+1}.json"
            dump_json(flp_json, {
                "instance": filename,
                "run_index": run_index + 1,
                **flp_solver.last_solution
            })

        logging.debug("FLP Results: %s", flp_results)
        od = len(flp_results[1])

        flp_execution_time = flp_results[5]

        ve_st = datetime.now()
        print("Running VROOM function")
        VROOM_solver = dataCvrp(ans, flp_results)
        VROOM_results = VROOM_solver.runVROOM()

        instance_stem = os.path.splitext(filename)[0]
        if getattr(VROOM_solver, "last_solution_vroom", None):
            vroom_json = SOLN_DIR / instance_stem / "VROOM" / f"run_{run_index+1}.json"
            dump_json(vroom_json, {
                "instance": filename,
                "run_index": run_index + 1,
                **VROOM_solver.last_solution_vroom
            })
            print(f"[saved] {vroom_json}")
        else:
            print(f"[WARN] No VROOM JSON payload for {filename}, run {run_index+1}")


        logging.info("VRP Results: %s", VROOM_results)
        ve_ed = datetime.now()
        tot_ve_exec = (ve_ed - ve_st).total_seconds()
        ve_exec = tot_ve_exec / od if od != 0 else 0

        bks = bks_dict.get(filename)
        VROOM_total_lrp_cost = VROOM_results[0]
        if bks is not None and bks != 0:
            gap = ((VROOM_total_lrp_cost - bks) / bks) * 100
            gap = round(gap, 2)
        else:
            gap = None

        VROOM_VRP_COST = VROOM_results[1]
        if VROOM_VRP_COST is not None and VROOM_VRP_COST != 0:
            surr_gap = ((sur_cost - VROOM_VRP_COST) / VROOM_VRP_COST )*100
            surr_gap = round(surr_gap, 2)
        else:
            surr_gap = None

        flp_cost_list.append(flp_results[0])
        flp_execution_time_list.append(flp_execution_time)
        VROOM_total_vrp_cost_list.append(VROOM_results[1])
        VROOM_total_lrp_cost_list.append(VROOM_total_lrp_cost)
        VROOM_total_execution_time_list.append(tot_ve_exec)
        VROOM_execution_time_per_operation_list.append(ve_exec)
        gap_list.append(gap)
        sur_cost_list.append(sur_cost)
        gap_sugg_list.append(surr_gap)

        for handler in logging.root.handlers[:]:
            handler.close()
            logging.root.removeHandler(handler)

    avg_flp_cost = sum(flp_cost_list) / len(flp_cost_list)
    avg_flp_execution_time = sum(flp_execution_time_list) / len(flp_execution_time_list)
    avg_VROOM_total_vrp_cost = sum(VROOM_total_vrp_cost_list) / len(VROOM_total_vrp_cost_list)
    avg_VROOM_total_lrp_cost = sum(VROOM_total_lrp_cost_list) / len(VROOM_total_lrp_cost_list)
    avg_VROOM_total_execution_time = sum(VROOM_total_execution_time_list) / len(VROOM_total_execution_time_list)
    avg_VROOM_execution_time_per_operation = sum(VROOM_execution_time_per_operation_list) / len(VROOM_execution_time_per_operation_list)
    avg_sur_cost = sum(sur_cost_list) / len(sur_cost_list)

    valid_gaps = [g for g in gap_list if g is not None]
    avg_gap = (sum(valid_gaps) / len(valid_gaps)) if valid_gaps else "N/A"

    valid_sugg = [g for g in gap_sugg_list if g is not None]
    avg_surr_gap = (sum(valid_sugg) / len(valid_sugg)) if valid_sugg else "N/A"

    new_row = [
        os.path.basename(file_path),
        avg_flp_cost,
        avg_sur_cost,
        avg_flp_execution_time,
        avg_VROOM_total_vrp_cost,
        avg_VROOM_total_lrp_cost,
        avg_VROOM_total_execution_time,
        avg_VROOM_execution_time_per_operation,
        avg_gap,
        avg_surr_gap
    ]

    worksheet.append(new_row)

    workbook.save(existing_excel_file)

    processed_instances.add(filename)